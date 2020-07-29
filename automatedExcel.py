import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
for row in range(2, sheet.max_row + 1):
    incorrectValue = sheet.cell(row, 3)
    correctValue = incorrectValue.value * 0.9
    correctCell = sheet.cell(row, 4)
    correctCell.value = correctValue
wb.save('transactions2.xlsx')

values = Reference(sheet, min_row=2, max_row=4, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'f3')
